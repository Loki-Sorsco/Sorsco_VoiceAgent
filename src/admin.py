"""Backend for the dashboard: clients, store webhooks, call queue, history.

The UI itself lives in src/admin_ui.py (served at /admin). Registered onto the
Pipecat runner's FastAPI app by run_web.py, so one domain serves everything.
"""

import base64
import hashlib
import hmac
import json
import os
import random
from pathlib import Path

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, Response
from loguru import logger

from src.admin_ui import ADMIN_PAGE
from src.call_events import count_events, read_events, read_history
from src.store import get_task, list_tasks, set_active_task, update_task
from src.triggers import (
    DEFAULT_TRIGGERS,
    evaluate,
    queue_manual_call,
    woocommerce_to_shopify,
)

ROOT = Path(__file__).resolve().parent.parent
CLIENTS_DIR = ROOT / "clients"
ACTIVE_FILE = ROOT / "data" / "active_client.txt"
ORDERS_FILE = ROOT / "data" / "orders.json"


def get_active_client_id(default: str = "hotel_sunrise") -> str:
    if ACTIVE_FILE.exists():
        active = ACTIVE_FILE.read_text(encoding="utf-8").strip()
        if active and (CLIENTS_DIR / f"{active}.json").exists():
            return active
    return default


def _load_client(client_id: str) -> dict:
    path = CLIENTS_DIR / f"{client_id}.json"
    if not path.exists():
        raise HTTPException(404, f"No client '{client_id}'")
    return json.loads(path.read_text(encoding="utf-8"))


def register_admin(app: FastAPI):
    # ------------------------------------------------------------- clients

    @app.get("/api/clients")
    def list_clients():
        out = []
        for p in sorted(CLIENTS_DIR.glob("*.json")):
            try:
                cfg = json.loads(p.read_text(encoding="utf-8"))
                out.append(
                    {
                        "client_id": cfg.get("client_id", p.stem),
                        "business_name": cfg.get("business_name", p.stem),
                        "agent_name": cfg.get("agent_name", ""),
                        "default_language": cfg.get("default_language", "hi-IN"),
                        "tts_voice": cfg.get("tts_voice", "priya"),
                        "shopify_connected": bool(
                            (cfg.get("shopify") or {}).get("access_token")
                        ),
                    }
                )
            except json.JSONDecodeError:
                continue
        return {"clients": out, "active": get_active_client_id()}

    @app.get("/api/clients/{client_id}")
    def get_client(client_id: str):
        cfg = _load_client(client_id)
        cfg.setdefault("triggers", DEFAULT_TRIGGERS)
        cfg.setdefault("shopify", {"domain": "", "access_token": "", "webhook_secret": ""})
        return cfg

    @app.post("/api/clients/{client_id}")
    def save_client(client_id: str, cfg: dict):
        client_id = client_id.strip().lower().replace(" ", "_").replace("-", "_")
        if not client_id.replace("_", "").isalnum():
            raise HTTPException(400, "Agent ID must be letters, numbers, underscores")
        required = ["business_name", "agent_name", "persona", "knowledge"]
        missing = [f for f in required if not cfg.get(f)]
        if missing:
            raise HTTPException(400, f"Missing fields: {', '.join(missing)}")
        cfg["client_id"] = client_id
        cfg.setdefault("default_language", "hi-IN")
        cfg.setdefault("supported_languages", ["hi-IN", "en-IN"])
        cfg.setdefault("tts_voice", "priya")
        CLIENTS_DIR.mkdir(exist_ok=True)
        path = CLIENTS_DIR / f"{client_id}.json"
        path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"saved": client_id}

    @app.post("/api/active-client/{client_id}")
    def set_active(client_id: str):
        _load_client(client_id)
        ACTIVE_FILE.parent.mkdir(exist_ok=True)
        ACTIVE_FILE.write_text(client_id, encoding="utf-8")
        return {"active": client_id}

    # -------------------------------------------------------- voice preview

    @app.get("/api/voice-preview/{voice}")
    def voice_preview(voice: str, lang: str = "hi-IN"):
        """Short Sarvam TTS sample so users can audition voices."""
        try:
            from sarvamai import SarvamAI

            client = SarvamAI(api_subscription_key=os.environ["SARVAM_API_KEY"])
            r = client.text_to_speech.convert(
                text="Namaste! Main aapki AI agent hoon. Aapki kya madad kar sakti hoon?",
                target_language_code=lang,
                model="bulbul:v3",
                speaker=voice,
            )
            audio = base64.b64decode(r.audios[0])
            return Response(content=audio, media_type="audio/wav")
        except Exception as e:
            logger.warning(f"Voice preview failed: {e}")
            raise HTTPException(500, f"Preview failed: {e}")

    # ----------------------------------------------------- shopify webhooks

    @app.post("/webhooks/shopify/{client_id}")
    async def shopify_webhook(client_id: str, request: Request):
        cfg = _load_client(client_id)
        raw = await request.body()

        secret = (cfg.get("shopify") or {}).get("webhook_secret", "")
        if secret:
            sent = request.headers.get("X-Shopify-Hmac-Sha256", "")
            digest = base64.b64encode(
                hmac.new(secret.encode(), raw, hashlib.sha256).digest()
            ).decode()
            if not hmac.compare_digest(digest, sent):
                logger.warning(f"Shopify webhook HMAC mismatch for {client_id}")
                raise HTTPException(401, "HMAC verification failed")

        topic = request.headers.get("X-Shopify-Topic", "orders/create")
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            raise HTTPException(400, "Invalid JSON")

        task = evaluate(topic, payload, cfg)
        return {"received": topic, "call_queued": task["task_id"] if task else None}

    @app.post("/api/simulate-order/{client_id}")
    def simulate_order(client_id: str, body: dict):
        """Free demo mode: fake a Shopify orders/create webhook."""
        cfg = _load_client(client_id)
        kind = body.get("type", "cod")
        n = random.randint(1000, 9999)
        payload = {
            "id": n,
            "name": f"#{n}",
            "total_price": str(random.choice([799, 1499, 2499, 3999])),
            "currency": "INR",
            "financial_status": "pending",
            "payment_gateway_names": (
                ["Cash on Delivery (COD)"] if kind == "cod" else ["razorpay"]
            ),
            "customer": {"first_name": "Rahul", "last_name": "Sharma", "phone": "+919876501234"},
            "shipping_address": {
                "address1": "42 MG Road",
                "city": "Jaipur",
                "province": "Rajasthan",
                "zip": "302001",
                "phone": "+919876501234",
            },
            "line_items": [
                {
                    "title": random.choice(
                        ["Cotton Kurta (L)", "Wireless Earbuds", "Steel Water Bottle 1L"]
                    ),
                    "quantity": 1,
                    "price": "1499",
                }
            ],
        }
        task = evaluate("orders/create", payload, cfg)
        if not task:
            return {"queued": None, "note": "No trigger fired — check trigger settings."}
        return {"queued": task["task_id"], "order": payload["name"]}

    # ------------------------------------------------------------ call queue

    @app.get("/api/queue")
    def queue():
        return {"tasks": list(reversed(list_tasks()))[:50]}

    @app.post("/api/queue/{task_id}/take")
    def take_task(task_id: str):
        task = get_task(task_id)
        if not task:
            raise HTTPException(404, "No such task")
        ACTIVE_FILE.parent.mkdir(exist_ok=True)
        ACTIVE_FILE.write_text(task["client_id"], encoding="utf-8")
        set_active_task(task_id)
        return {"armed": task_id}

    @app.post("/api/queue/{task_id}/dismiss")
    def dismiss_task(task_id: str):
        update_task(task_id, status="done", outcome="dismissed")
        return {"dismissed": task_id}

    # ------------------------------------------- more integrations & extras

    @app.post("/webhooks/woocommerce/{client_id}")
    async def woocommerce_webhook(client_id: str, request: Request):
        cfg = _load_client(client_id)
        try:
            payload = json.loads(await request.body())
        except json.JSONDecodeError:
            raise HTTPException(400, "Invalid JSON")
        task = evaluate("orders/create", woocommerce_to_shopify(payload), cfg)
        return {"received": "woocommerce order", "call_queued": task["task_id"] if task else None}

    @app.post("/webhooks/generic/{client_id}")
    async def generic_webhook(client_id: str, body: dict):
        """Universal trigger: any system (Zapier, Sheets, your CRM) queues a call.

        Body: {"name": "...", "phone": "...", "purpose": "why the agent should call"}
        """
        cfg = _load_client(client_id)
        name = body.get("name") or body.get("customer_name") or "the customer"
        phone = body.get("phone") or body.get("customer_phone") or ""
        purpose = body.get("purpose") or body.get("reason") or "Follow up with this customer"
        task = queue_manual_call(cfg, name, phone, purpose)
        return {"call_queued": task["task_id"]}

    @app.post("/api/campaign/{client_id}")
    def campaign(client_id: str, body: dict):
        """Bulk calls: entries [{name, phone}] + one purpose for all of them."""
        cfg = _load_client(client_id)
        purpose = (body.get("purpose") or "").strip()
        entries = body.get("entries") or []
        if not purpose:
            raise HTTPException(400, "Purpose is required")
        if not entries:
            raise HTTPException(400, "No contacts given")
        tasks = [
            queue_manual_call(cfg, e.get("name", "the customer"), e.get("phone", ""), purpose)
            for e in entries[:200]
        ]
        return {"queued": len(tasks)}

    @app.get("/api/shopify/test/{client_id}")
    async def shopify_test(client_id: str):
        cfg = _load_client(client_id)
        shop = cfg.get("shopify") or {}
        domain, token = shop.get("domain"), shop.get("access_token")
        if not (domain and token):
            raise HTTPException(400, "Store domain and access token are not set")
        try:
            import aiohttp

            async with aiohttp.ClientSession() as session:
                async with session.get(
                    f"https://{domain}/admin/api/2024-10/shop.json",
                    headers={"X-Shopify-Access-Token": token},
                    timeout=aiohttp.ClientTimeout(total=10),
                ) as r:
                    if r.status == 200:
                        data = await r.json()
                        return {"ok": True, "shop": data.get("shop", {}).get("name", domain)}
                    raise HTTPException(400, f"Shopify answered HTTP {r.status} — check the token")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Could not reach the store: {e}")

    @app.post("/api/chat-test/{client_id}")
    def chat_test(client_id: str, body: dict):
        """Text-only agent test for the console (no voice, no tools)."""
        from src.config_loader import build_system_prompt
        from src.llm_factory import chat_complete

        cfg = _load_client(client_id)
        messages = (body.get("messages") or [])[-12:]
        try:
            reply = chat_complete(build_system_prompt(cfg), messages)
            return {"reply": reply}
        except Exception as e:
            logger.warning(f"chat-test failed: {e}")
            raise HTTPException(500, str(e))

    @app.post("/api/queue/{task_id}/requeue")
    def requeue_task(task_id: str):
        task = get_task(task_id)
        if not task:
            raise HTTPException(404, "No such task")
        update_task(task_id, status="queued", outcome=None)
        return {"requeued": task_id}

    @app.get("/api/events")
    def live_events(since: int = 0):
        return {"events": read_events(since), "count": count_events()}

    # ------------------------------------------------------- real telephony

    @app.get("/api/telephony/status")
    def get_telephony_status():
        from src.telephony import telephony_status

        return telephony_status()

    @app.post("/api/queue/{task_id}/dial")
    async def dial_task(task_id: str):
        from src.telephony import place_call

        task = get_task(task_id)
        if not task:
            raise HTTPException(404, "No such task")
        try:
            result = await place_call(task)
        except ValueError as e:
            raise HTTPException(400, str(e))
        update_task(task_id, status="dialing")
        return {"dialing": task_id, "call_sid": result["call_sid"]}

    @app.post("/api/queue/dial-all")
    async def dial_all():
        from src.telephony import place_call

        dialed, errors = [], []
        for task in list_tasks():
            if task["status"] != "queued":
                continue
            try:
                await place_call(task)
                update_task(task["task_id"], status="dialing")
                dialed.append(task["task_id"])
            except ValueError as e:
                errors.append(f"{task.get('customer_name','?')}: {e}")
        return {"dialed": len(dialed), "errors": errors[:5]}

    @app.post("/api/shopify/pull/{client_id}")
    async def shopify_pull(client_id: str, body: dict):
        """Fetch orders from the connected Shopify store and queue calls.

        body.kind: "pending" (payment pending) or "tag" (body.tag matches order tags)
        """
        cfg = _load_client(client_id)
        shop = cfg.get("shopify") or {}
        domain, token = shop.get("domain"), shop.get("access_token")
        if not (domain and token):
            raise HTTPException(400, "Connect the Shopify store first (domain + token)")

        kind = body.get("kind", "pending")
        tag = (body.get("tag") or "").strip().lower()
        params = "status=any&limit=50"
        if kind == "pending":
            params += "&financial_status=pending"

        import aiohttp

        from src.triggers import FLOW_DESCRIPTIONS, normalize_shopify_order
        from src.store import add_task, save_order

        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    f"https://{domain}/admin/api/2024-10/orders.json?{params}",
                    headers={"X-Shopify-Access-Token": token},
                    timeout=aiohttp.ClientTimeout(total=15),
                ) as r:
                    if r.status != 200:
                        raise HTTPException(400, f"Shopify answered HTTP {r.status}")
                    orders = (await r.json()).get("orders", [])
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Could not reach the store: {e}")

        queued = 0
        existing = {t["order_id"] for t in list_tasks() if t["status"] in ("queued", "dialing")}
        for o in orders:
            if kind == "tag" and tag not in (o.get("tags", "") or "").lower():
                continue
            order = normalize_shopify_order(o, client_id)
            if str(order["order_id"]) in existing:
                continue
            save_order(order)
            reason = (
                FLOW_DESCRIPTIONS["pending_payment"]
                if kind == "pending"
                else f"Follow up about their order (tagged '{tag}')"
            )
            add_task(client_id, order, reason, "pending_payment" if kind == "pending" else "campaign")
            queued += 1
        return {"queued": queued, "scanned": len(orders)}

    # ------------------------------------------------------ history & stats

    @app.get("/api/history")
    def history():
        from src.analysis import analyze_call
        from src.call_events import read_history_raw, rewrite_history

        records = read_history_raw()
        # Lazy backfill: analyze a few unanalyzed calls per request.
        todo = [r for r in records if "analysis" not in r][:5]
        if todo:
            for r in todo:
                r["analysis"] = analyze_call(r)
            rewrite_history(records)
        return {"calls": records[-100:][::-1]}

    @app.get("/api/history/export")
    def history_export():
        from openpyxl import Workbook
        from openpyxl.styles import Font

        from src.call_events import read_history_raw

        wb = Workbook()
        ws = wb.active
        ws.title = "Calls"
        headers = [
            "Started", "Agent", "Business", "Call type", "Outcome", "Duration (s)",
            "Turns", "Caller language", "Category", "Issue faced", "Summary",
            "Resolution", "Transcript",
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in read_history_raw():
            a = r.get("analysis") or {}
            transcript = " | ".join(
                f"{'C' if e.get('type') == 'user' else 'A'}: {e.get('text', '')}"
                for e in (r.get("transcript") or [])
                if e.get("type") in ("user", "assistant")
            )
            ws.append([
                r.get("started", ""), r.get("agent", ""), r.get("client", ""),
                (r.get("kind") or "").replace("_", " "), r.get("outcome") or "",
                r.get("duration_s", 0), r.get("turns", 0), a.get("language", ""),
                a.get("category", ""), a.get("issue", ""), a.get("summary", ""),
                a.get("resolution", ""), transcript[:3000],
            ])
        widths = [10, 12, 22, 16, 14, 12, 8, 14, 18, 40, 50, 16, 80]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        import io

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="call-report.xlsx"'},
        )

    @app.get("/api/stats")
    def stats():
        calls = read_history(1000)
        orders = {}
        if ORDERS_FILE.exists():
            try:
                orders = json.loads(ORDERS_FILE.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                pass
        confirmed_ids = [
            c["order_id"] for c in calls if c.get("outcome") == "confirmed" and c.get("order_id")
        ]
        revenue = sum(
            float(orders.get(str(oid), {}).get("total") or 0) for oid in set(confirmed_ids)
        )
        return {
            "calls_total": len(calls),
            "order_calls": sum(1 for c in calls if c.get("kind") != "inbound"),
            "confirmed": sum(1 for c in calls if c.get("outcome") == "confirmed"),
            "cancelled": sum(1 for c in calls if c.get("outcome") == "cancelled"),
            "revenue_confirmed": int(revenue),
            "minutes": round(sum(c.get("duration_s", 0) for c in calls) / 60, 1),
        }

    @app.get("/admin", response_class=HTMLResponse)
    def admin_page():
        return ADMIN_PAGE
